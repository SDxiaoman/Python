#-*- coding:utf-8 -*-
#"""
# Author:manjianchao 
# created time:2018-08-28
# """

import os
from openpyxl import load_workbook
from openpyxl import Workbook

#file_path='/home/yaoyijiang/manjianchao/Python/Code/files/test.xlsx'
file_path= "C:\\Users\\lsp\\Desktop\\2018年7月份车辆进场记录.xlsx"

#根据该关键字确定需要哪一列的数据
g_column_index_key='入场车牌号'
#需要从大表中提取的数据存放在该sheet下
g_template_sheet_name='车牌'
#待
g_source_sheet_name='inout'

#------------------------初始化-----------------------
def init():
    """
    加载原始Excel文件并获取关键信息对应的Sheet
    """
    print('init() called.')
    global g_excel
    global g_sheet
    global g_rows
    global g_columns    
    #加载Excel文件
    g_excel=load_workbook(file_path)
    print('加载Excel文件......')
    #加载对应的Sheet表格
    g_sheet=g_excel.get_sheet_by_name(g_template_sheet_name)
    print('获取sheet......')
    #获取最大的行数和列数
    g_rows=g_sheet.max_row
    print('rows=',g_rows)
    g_columns=g_sheet.max_column
    print('columns=',g_columns)

#--------------获取要过滤的车牌号--------------
def get_index_key():
    print('get_index_key() called.')
    for i in range(g_rows):
        for j in range(g_columns):
            #i,j的默认值是0开始，但是Excel的单元格从1开始
            cell_value=g_sheet.cell(row=i+1,column=j+1).value
            if cell_value == g_column_index_key:
                m_car_number_index_colume = j+1
                print('car number index=', m_car_number_index_colume)
                break
    global g_car_number_index_key=[]
    for k in range(2,g_rows):
        g_car_number_index_key.append(get_cell_value(g_sheet,k,m_car_number_index_colume))
    
    i=0
    for i in range(len(g_car_number_index_key)):
        print(g_car_number_index_key[i])


def filter_date_from_source_sheet():
    print('filter_date_from_source_sheet() called.')
    global g_result_sheet
    g_result_sheet=g_excel.create_sheet("result")

    i=j=0
    for i in range(g_source_sheet_rows):
        for j in range(g_source_sheet_columns):
            cell_value=g_sheet.cell(row=i+1,column=j+1).value
            if cell_value == g_column_index_key:
                car_number_index_colume = j+1
                print('car number index=', car_number_index_colume)
                break
    j=k=0
    for k in range(len(g_source_sheet_rows)):
        #读取到匹配的key
        if get_cell_value(g_source_sheet,k=1,car_number_index_colume) in g_car_number_index_key:
            #将该行写入一个新的sheet
            for j in range(len(g_source_sheet_columns)):
                print(get_cell_value(g_source_sheet,k+1,j+1))
                g_result_sheet.cell(row=k+1,column=j+1).value=get_cell_value(g_source_sheet,k+1,j+1)
        else:
            continue
    
    g_result_sheet.save(file_path)

  
def get_source_sheet():
    global g_source_sheet
    global g_source_sheet_rows
    global g_source_sheet_columns
    g_source_sheet=g_excel.get_sheet_by_name(g_source_sheet_name)
    g_source_sheet_rows=g_source_sheet.max_row
    g_source_sheet_columns=g_source_sheet.max_column
    
def get_cell_value(sheet,r,c):
    print('get_cell_value() called.')
    return sheet.cell(row=r,column=c).value


if __name__=='__main__':
    init()
    get_index_key()
    get_source_sheet()
    filter_date_from_source_sheet()
    
    